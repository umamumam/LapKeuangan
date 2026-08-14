import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Transaksi Supplier"

ws.views.sheetView[0].showGridLines = True

# Title Header
ws.merge_cells("A1:H1")
ws["A1"] = "LAPORAN TRANSAKSI SUPPLIER (BOS SOHEH)"
ws["A1"].font = Font(name="Segoe UI", size=15, bold=True, color="1A237E")
ws["A1"].alignment = Alignment(vertical="center")

ws.merge_cells("A2:H2")
ws["A2"] = "Format Otomatis Lusin & Potong, Total Jumlah, dan Sisa Tagihan"
ws["A2"].font = Font(name="Segoe UI", size=10, italic=True, color="555555")

headers = ["Tanggal", "Lsn", "Ptg", "Nama Barang", "@Harga (Lsn)", "Jumlah", "TF (Bayar)", "Tagihan"]
header_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)

num_fmt = 'Rp #,##0'

# Header Row
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border
ws.row_dimensions[4].height = 28

# Row 5: Sisa Nota Sebelumnya (Saldo Awal)
ws.cell(row=5, column=1, value="-").alignment = Alignment(horizontal="center")
ws.cell(row=5, column=4, value="SISA NOTA SEBELUMNYA (SALDO AWAL)").font = Font(name="Segoe UI", bold=True, italic=True)
ws.cell(row=5, column=8, value=0)
ws.cell(row=5, column=8).number_format = num_fmt
ws.cell(row=5, column=8).font = Font(name="Segoe UI", bold=True, color="C62828")
for c in range(1, 9):
    ws.cell(row=5, column=c).fill = PatternFill(start_color="FFF3E0", fill_type="solid")
    ws.cell(row=5, column=c).border = thin_border

# Sample Data Rows
data = [
    ("01/04/2026", 38, 0, "KS 27/30 (brown/green)", 760000, 0),
    ("01/04/2026", 17, 0, "KS 35/38 SNOW WSK", 830000, 20000000),
    ("02/04/2026", 2, 6, "Kaos Polos Cotton (Contoh Mix Lsn & Ptg)", 600000, 0),
    ("03/04/2026", 0, 18, "Celana Pendek (Contoh Hanya Ptg / 1.5 Lsn)", 480000, 500000),
]

start_row = 6
for i, row_data in enumerate(data):
    r = start_row + i
    ws.cell(row=r, column=1, value=row_data[0]).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2, value=row_data[1]).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=3, value=row_data[2]).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=4, value=row_data[3])
    ws.cell(row=r, column=5, value=row_data[4]).number_format = num_fmt
    
    # Formula Jumlah
    ws.cell(row=r, column=6, value=f"=(B{r}*E{r})+(C{r}*(E{r}/12))").number_format = num_fmt
    ws.cell(row=r, column=6).font = Font(name="Segoe UI", bold=True)
    
    # TF
    ws.cell(row=r, column=7, value=row_data[5]).number_format = num_fmt
    if row_data[5] > 0:
        ws.cell(row=r, column=7).font = Font(name="Segoe UI", color="C62828", bold=True)
        
    # Formula Tagihan = Tagihan_Sebelumnya + Jumlah - TF
    ws.cell(row=r, column=8, value=f"=H{r-1}+F{r}-G{r}").number_format = num_fmt
    ws.cell(row=r, column=8).font = Font(name="Segoe UI", bold=True, color="1565C0")
    
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = thin_border

# Add empty rows for extra entries
for i in range(len(data), len(data) + 15):
    r = start_row + i
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=5).number_format = num_fmt
    
    ws.cell(row=r, column=6, value=f'=IF(OR(B{r}<>"",C{r}<>""), (N(B{r})*E{r})+(N(C{r})*(E{r}/12)), 0)').number_format = num_fmt
    ws.cell(row=r, column=7).number_format = num_fmt
    
    ws.cell(row=r, column=8, value=f'=IF(AND(B{r}="",C{r}="",G{r}=""), "", H{r-1}+F{r}-N(G{r}))').number_format = num_fmt
    
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = thin_border

col_widths = {'A': 14, 'B': 10, 'C': 10, 'D': 40, 'E': 16, 'F': 18, 'G': 18, 'H': 20}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

output_path = "e:/Projek_Laporan/lapkeuangan/Template_Transaksi_Supplier.xlsx"
wb.save(output_path)
print(f"Excel generated successfully at {output_path}")
